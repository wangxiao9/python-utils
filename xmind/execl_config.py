__author__ = 'wangxiao'

import openpyxl


class ExcelConfig:
    ID = 1
    PATH = 2
    PRIORITY = 3
    TAG = 4
    SUMMARY = 5
    ACTION = 6
    RESULT = 7
    DESCRIPTION = 8

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sh = self.wb.create_sheet('test_case')

    def init_execl(self):
        self.sh.cell(row=1, column=self.ID, value='Test ID')
        self.sh.cell(row=1, column=self.PATH, value='Test Repository Path')
        self.sh.cell(row=1, column=self.PRIORITY, value='Priority')
        self.sh.cell(row=1, column=self.TAG, value='Tag')
        self.sh.cell(row=1, column=self.SUMMARY, value='Summary')
        self.sh.cell(row=1, column=self.ACTION, value='Action')
        self.sh.cell(row=1, column=self.RESULT, value='Result')
        self.sh.cell(row=1, column=self.DESCRIPTION, value='Description')

    def write_cases(self, i, *args, **kwargs):
        self.sh.cell(row=i, column=self.ID, value=str(kwargs['index']))
        self.sh.cell(row=i, column=self.PATH, value=kwargs['path'])
        self.sh.cell(row=i, column=self.PRIORITY, value=kwargs['priority'])
        self.sh.cell(row=i, column=self.TAG, value=kwargs['tag'])
        self.sh.cell(row=i, column=self.SUMMARY, value=kwargs['summary'])
        self.sh.cell(row=i, column=self.DESCRIPTION, value=kwargs['des'])

        self.sh.cell(row=i, column=self.ACTION, value=args[0][0])
        self.sh.cell(row=i, column=self.RESULT, value=args[0][1])


    def run(self, i, *args, **kwargs):
        self.init_execl()
        self.write_cases(i, *args, **kwargs)
        self.wb.save('case.xlsx')


if __name__ == '__main__':
    ExcelConfig().write_value(2, 'test_001')
